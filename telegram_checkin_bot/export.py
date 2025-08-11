import os
import re
import pandas as pd
import pytz
import logging
from datetime import datetime
from config import DATA_DIR, DATABASE_URL, BEIJING_TZ
import cloudinary
import cloudinary.uploader
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from shift_manager import get_shift_times_short
from sqlalchemy import create_engine
from db_pg import get_conn 


# ===========================
# 基础配置
# ===========================
logging.basicConfig(level=logging.INFO, format="[%(asctime)s] %(levelname)s: %(message)s")

# ===========================
# 文件名安全化（去除非法字符）
# ===========================
def safe_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", str(name))

# ===========================
# 上传文件到 Cloudinary
# ===========================
def upload_to_cloudinary(file_path: str) -> str | None:
    try:
        result = cloudinary.uploader.upload(
            file_path,
            resource_type="raw",
            folder="telegram_exports",
            public_id=os.path.splitext(os.path.basename(file_path))[0]
        )
        return result.get("secure_url")
    except Exception as e:
        logging.error(f"❌ Cloudinary 上传失败: {e}")
        return None

# ===========================
# 读取数据库数据到 DataFrame
# ===========================
def _fetch_data(start_datetime: datetime, end_datetime: datetime) -> pd.DataFrame:
    try:
        engine = create_engine(DATABASE_URL)
        query = """
        SELECT username, name, content, timestamp, keyword, shift 
        FROM messages 
        WHERE timestamp BETWEEN %(start)s AND %(end)s
        """
        params = {
            "start": start_datetime.astimezone(pytz.UTC),
            "end": end_datetime.astimezone(pytz.UTC)
        }
        # 分块读取（避免大数据内存溢出）
        df_iter = pd.read_sql_query(query, engine, params=params, chunksize=50000)
        df = pd.concat(df_iter, ignore_index=True)
        logging.info(f"✅ 数据读取完成，共 {len(df)} 条记录")
    except Exception as e:
        logging.error(f"❌ 无法连接数据库或读取数据: {e}")
        return pd.DataFrame()

    if df.empty:
        return df

    # 时间转为北京时区
    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce", utc=True).dt.tz_convert(BEIJING_TZ)
    df = df.dropna(subset=["timestamp"]).copy()
    return df

# ===========================
# Excel 内标记迟到/早退/补卡
# ===========================
def _mark_late_early(excel_path: str):
    wb = load_workbook(excel_path)
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # 浅红色填充（异常）
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 浅黄色填充（补卡）

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2):  # 跳过表头
            shift_cell, time_cell, keyword_cell = row[3], row[1], row[2]

            if not shift_cell.value or not time_cell.value:
                continue

            shift_text = str(shift_cell.value).strip()
            shift_name = re.split(r'[（(]', shift_text)[0]  # 班次名（去除括号）

            # 解析时间（兼容 Excel datetime 和字符串）
            if isinstance(time_cell.value, datetime):
                dt = time_cell.value
            else:
                try:
                    dt = datetime.strptime(str(time_cell.value), "%Y-%m-%d %H:%M:%S")
                except Exception:
                    continue

            # 1️⃣ 补卡标记
            if "补卡" in shift_text:
                time_cell.fill = fill_yellow
                shift_cell.fill = fill_yellow
                if "（补卡）" not in shift_text:
                    shift_cell.value = f"{shift_text}（补卡）"
                continue

            # 2️⃣ 迟到/早退判定
            if shift_name in get_shift_times_short():
                start_time, end_time = get_shift_times_short()[shift_name]

                # ---- 迟到 ----
                if keyword_cell.value == "#上班打卡" and dt.time() > start_time:
                    time_cell.fill = fill_red
                    shift_cell.fill = fill_red
                    if "（迟到）" not in shift_text:
                        shift_cell.value = f"{shift_text}（迟到）"

                # ---- 早退 ----
                elif keyword_cell.value == "#下班打卡":
                    if shift_name == "I班":
                        if dt.hour == 0:  # I班次日 00:00 正常
                            continue
                        elif 15 <= dt.hour <= 23:  # 当天提早下班
                            time_cell.fill = fill_red
                            shift_cell.fill = fill_red
                            if "（早退）" not in shift_text:
                                shift_cell.value = f"{shift_text}（早退）"
                    else:
                        if 0 <= dt.hour <= 1:  # 跨天凌晨下班正常
                            continue
                        if dt.time() < end_time:  # 提前下班
                            time_cell.fill = fill_red
                            shift_cell.fill = fill_red
                            if "（早退）" not in shift_text:
                                shift_cell.value = f"{shift_text}（早退）"

    wb.save(excel_path)
    
# 获取所有用户姓名
def get_all_user_names():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT name FROM users;")
            return [row[0] for row in cur.fetchall()]
            
# ===========================
# 导出打卡记录 Excel
# ===========================
def export_excel(start_datetime: datetime, end_datetime: datetime):
    df = _fetch_data(start_datetime, end_datetime)
    if df.empty:
        logging.warning("⚠️ 指定日期内没有数据")
        return None

    # 去掉时区
    if pd.api.types.is_datetime64_any_dtype(df["timestamp"]):
        try:
            df["timestamp"] = df["timestamp"].dt.tz_localize(None)
        except AttributeError:
            pass

    df["date"] = df["timestamp"].dt.strftime("%Y-%m-%d")
    start_str = start_datetime.strftime("%Y-%m-%d")
    end_str = (end_datetime - pd.Timedelta(seconds=1)).strftime("%Y-%m-%d")

    export_dir = os.path.join(DATA_DIR, f"excel_{start_str}_{end_str}")
    os.makedirs(export_dir, exist_ok=True)
    excel_path = os.path.join(export_dir, f"打卡记录_{start_str}_{end_str}.xlsx")

    all_user_names = get_all_user_names()

    def format_shift(shift):
        if pd.isna(shift):
            return shift
        shift_text = str(shift)
        if re.search(r'（\d{2}:\d{2}-\d{2}:\d{2}）', shift_text):
            return shift_text
        shift_name = shift_text.split("（")[0]
        if shift_name in get_shift_times_short():
            start, end = get_shift_times_short()[shift_name]
            return f"{shift_text}（{start.strftime('%H:%M')}-{end.strftime('%H:%M')}）"
        return shift_text

    # 统计未打卡
    missed_days_count = {u: 0 for u in all_user_names}

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for day, group_df in df.groupby("date"):
            checked_users = set(group_df["name"].unique())
            missed_users = [u for u in all_user_names if u not in checked_users]

            group_df = group_df.copy()
            group_df["remark"] = ""

            for u in missed_users:
                missed_days_count[u] += 1

            if missed_users:
                missed_df = pd.DataFrame({
                    "name": missed_users,
                    "timestamp": pd.NaT,
                    "keyword": None,
                    "shift": None,
                    "remark": "当天未打卡"
                })
                group_df = pd.concat([group_df, missed_df], ignore_index=True)

            group_df = group_df.sort_values("timestamp", na_position="last")
            slim_df = group_df[["name", "timestamp", "keyword", "shift", "remark"]].copy()
            slim_df.columns = ["姓名", "打卡时间", "关键词", "班次", "备注"]

            slim_df["打卡时间"] = pd.to_datetime(slim_df["打卡时间"], errors="coerce").dt.tz_localize(None)
            slim_df["班次"] = slim_df["班次"].apply(format_shift)
            slim_df.to_excel(writer, sheet_name=day[:31], index=False)
            
    # 标记迟到/早退/补卡
    _mark_late_early(excel_path)
    
    # 打开写好的文件
    wb = load_workbook(excel_path)

    # 每日明细表格样式
    red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # 淡红
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 淡黄
    blue_fill_light = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # 淡蓝

    for sheet in wb.worksheets:
        if sheet.title == "统计":
            continue
        for row in sheet.iter_rows(min_row=2):
            remark_val = str(row[4].value or "")
            if "迟到" in remark_val or "早退" in remark_val:
                for cell in row:
                    cell.fill = red_fill
            elif "补卡" in remark_val:
                for cell in row:
                    cell.fill = yellow_fill
            elif "未打卡" in remark_val:
                for cell in row:
                    cell.fill = blue_fill_light

    # 生成统计数据
    stats = []
    for sheet in wb.worksheets:
        if sheet.title == "统计":
            continue
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, _, keyword, shift_text, remark = row
            if not name:
                continue
            if remark == "当天未打卡":
                continue
            elif "补卡" in str(shift_text):
                status = "补卡"
            elif "迟到" in str(shift_text) or "早退" in str(shift_text):
                status = "迟到/早退"
            else:
                status = "正常"
            stats.append({"姓名": name, "状态": status})

    stats_df = pd.DataFrame(stats)
    if not stats_df.empty:
        summary_df = stats_df.groupby(["姓名", "状态"]).size().unstack(fill_value=0).reset_index()
    else:
        summary_df = pd.DataFrame(columns=["姓名", "正常", "迟到/早退", "补卡"])

    # 确保所有用户都在统计表
    for user in all_user_names:
        if user not in summary_df["姓名"].values:
            summary_df = pd.concat([
                summary_df,
                pd.DataFrame([{"姓名": user}])
            ], ignore_index=True)

    for col in ["正常", "迟到/早退", "补卡"]:
        if col not in summary_df.columns:
            summary_df[col] = 0
    summary_df = summary_df.fillna(0).astype({"正常": int, "迟到/早退": int, "补卡": int})

    # 未打卡次数
    summary_df["未打卡"] = summary_df["姓名"].map(missed_days_count)
    summary_df["异常总数"] = summary_df["迟到/早退"] + summary_df["补卡"]

    summary_df = summary_df[["姓名", "正常", "未打卡", "迟到/早退", "补卡", "异常总数"]]
    summary_df = summary_df.sort_values(by="正常", ascending=False)

    # 写入统计表
    stats_sheet = wb.create_sheet("统计", 0)
    headers = ["姓名", "正常打卡", "未打卡", "迟到/早退", "补卡", "异常总数"]
    for r_idx, row in enumerate([headers] + summary_df.values.tolist(), 1):
        for c_idx, value in enumerate(row, 1):
            stats_sheet.cell(row=r_idx, column=c_idx, value=value)

    stats_sheet.freeze_panes = "A2"

    # 样式
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    for cell in stats_sheet[1]:
        cell.font = header_font
        cell.alignment = center_align

    # 给 迟到/早退、补卡、异常总数 列加淡蓝色背景
    for row in stats_sheet.iter_rows(min_row=2):
        for col_idx in [4, 5, 6]:  # 第4,5,6列
            row[col_idx - 1].fill = blue_fill

    # 调整所有表样式
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = center_align
        for col in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = max_length + 2
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(excel_path)
    logging.info(f"✅ Excel 导出完成: {excel_path}")
    return excel_path
