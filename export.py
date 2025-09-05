import os
import re
import pandas as pd
import pytz
import logging
from datetime import datetime, timedelta
from config import DATA_DIR, DATABASE_URL, BEIJING_TZ
import cloudinary
import cloudinary.uploader
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from shift_manager import get_shift_times_short
from sqlalchemy import create_engine
from db_pg import get_conn 


# ===========================
# 基础配置
# ===========================
logging.basicConfig(level=logging.INFO, format="[%(asctime)s] %(levelname)s: %(message)s")

# ===========================
# 文件名安全化（去除非法字符）
# ===========================
def safe_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", str(name))

# ===========================
# 上传文件到 Cloudinary
# ===========================
def upload_to_cloudinary(file_path: str) -> str | None:
    try:
        result = cloudinary.uploader.upload(
            file_path,
            resource_type="raw",
            folder="telegram_exports",
            public_id=os.path.splitext(os.path.basename(file_path))[0]
        )
        return result.get("secure_url")
    except Exception as e:
        logging.error(f"❌ Cloudinary 上传失败: {e}")
        return None

# ===========================
# 读取数据库数据到 DataFrame
# ===========================
def _fetch_data(start_datetime: datetime, end_datetime: datetime) -> pd.DataFrame:
    try:
        engine = create_engine(DATABASE_URL)
        query = """
        SELECT username, name, content, timestamp, keyword, shift 
        FROM messages 
        WHERE timestamp BETWEEN %(start)s AND %(end)s
        """
        params = {
            "start": start_datetime.astimezone(pytz.UTC),
            "end": end_datetime.astimezone(pytz.UTC)
        }
        # 分块读取（避免大数据内存溢出）
        df_iter = pd.read_sql_query(query, engine, params=params, chunksize=50000)
        df = pd.concat(df_iter, ignore_index=True)
        logging.info(f"✅ 数据读取完成，共 {len(df)} 条记录")
    except Exception as e:
        logging.error(f"❌ 无法连接数据库或读取数据: {e}")
        return pd.DataFrame()

    if df.empty:
        return df

    # 时间转为北京时区
    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce", utc=True).dt.tz_convert(BEIJING_TZ)
    df = df.dropna(subset=["timestamp"]).copy()
    return df

# 获取所有用户姓名
def get_all_user_names():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT name FROM users;")
            return [row[0] for row in cur.fetchall()]

# 导出打卡记录（修正版）
def export_excel(start_datetime: datetime, end_datetime: datetime):
    df = _fetch_data(start_datetime, end_datetime)
    if df.empty:
        logging.warning("⚠️ 指定日期内没有数据")
        export_dir = os.path.join(DATA_DIR, f"excel_{start_datetime:%Y-%m-%d}_{end_datetime - pd.Timedelta(seconds=1):%Y-%m-%d}")
        os.makedirs(export_dir, exist_ok=True)
        excel_path = os.path.join(export_dir, f"打卡记录_{start_datetime:%Y-%m-%d}_{end_datetime - pd.Timedelta(seconds=1):%Y-%m-%d}.xlsx")
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            pd.DataFrame(columns=["姓名", "打卡时间", "关键词", "班次", "备注"]).to_excel(writer, sheet_name="空表", index=False)
        return excel_path

    # ======================== 时间处理 ========================
    if pd.api.types.is_datetime64_any_dtype(df["timestamp"]):
        try:
            df["timestamp"] = df["timestamp"].dt.tz_localize(None)
        except AttributeError:
            pass

    df["date"] = df["timestamp"].dt.strftime("%Y-%m-%d")
    start_str = start_datetime.strftime("%Y-%m-%d")
    end_str = (end_datetime - pd.Timedelta(seconds=1)).strftime("%Y-%m-%d")

    export_dir = os.path.join(DATA_DIR, f"excel_{start_str}_{end_str}")
    os.makedirs(export_dir, exist_ok=True)
    excel_path = os.path.join(export_dir, f"打卡记录_{start_str}_{end_str}.xlsx")

    all_user_names = get_all_user_names()

    def format_shift(shift):
        if pd.isna(shift):
            return shift
        shift_text = str(shift)
        if re.search(r'（\d{2}:\d{2}-\d{2}:\d{2}）', shift_text):
            return shift_text
        shift_name = shift_text.split("（")[0]
        if shift_name in get_shift_times_short():
            start, end = get_shift_times_short()[shift_name]
            return f"{shift_text}（{start.strftime('%H:%M')}-{end.strftime('%H:%M')}）"
        return shift_text

    missed_days_count = {u: 0 for u in all_user_names}

    # 过滤掉当天 sheet 的 I班凌晨下班卡（次日）
    i_shift_mask = (
        (df["keyword"] == "#下班打卡") &
        (df["shift"].notna()) &
        (df["shift"].astype(str).str.startswith("I班")) &
        (df["timestamp"].dt.hour < 6)
    )
    cross_df = df[i_shift_mask].copy()
    df = df[~i_shift_mask]
    cross_df["remark"] = cross_df.get("remark", "") + "（次日）"
    cross_df["date"] = (cross_df["timestamp"] - pd.Timedelta(days=1)).dt.strftime("%Y-%m-%d")
    df = pd.concat([df, cross_df], ignore_index=True)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        sheet_written = False

        for day, group_df in sorted(df.groupby("date"), key=lambda x: x[0], reverse=True):
            group_df = group_df.copy()
            if "remark" not in group_df.columns:
                group_df["remark"] = ""

            # 当日已打上班 / 下班的用户
            checked_users = set(group_df.loc[group_df["keyword"] == "#上班打卡", "name"].unique())
            down_checked_users = set(group_df.loc[group_df["keyword"] == "#下班打卡", "name"].unique())  # ✅ 修复

            missed_users = []
            day_date = datetime.strptime(day, "%Y-%m-%d").date()

            for u in all_user_names:
                if u not in checked_users:
                    missed_users.append(u)
                    missed_days_count[u] += 1
                elif u not in down_checked_users:
                    # 有上班但没下班 → 未打下班卡
                    group_df = pd.concat([
                        group_df,
                        pd.DataFrame([{
                            "name": u,
                            "timestamp": pd.NaT,
                            "keyword": "#下班打卡",
                            "shift": None,
                            "remark": "未打下班卡"
                        }])
                    ], ignore_index=True)

            if missed_users:
                missed_df = pd.DataFrame({
                    "name": missed_users,
                    "timestamp": pd.NaT,
                    "keyword": None,
                    "shift": None,
                    "remark": "休息/缺勤"
                })
                group_df = pd.concat([group_df, missed_df], ignore_index=True)

            # ======================== 迟到/早退/补卡 ========================
            for idx, row in group_df.iterrows():
                shift_val = row["shift"]
                keyword = row["keyword"]
                ts = row["timestamp"]

                if not shift_val or pd.isna(ts):
                    continue

                shift_text = str(shift_val).strip()
                shift_name = re.split(r'[（(]', shift_text)[0]

                if "补卡" in shift_text:
                    group_df.at[idx, "remark"] = "补卡"
                    continue

                if shift_name in get_shift_times_short():
                    start_time, end_time = get_shift_times_short()[shift_name]
                    ts_time = ts.time()

                    if keyword == "#上班打卡" and ts_time > start_time:
                        group_df.at[idx, "remark"] = "迟到"
                    elif keyword == "#下班打卡":
                        if shift_name == "I班":
                            if not (ts.hour == 0):
                                if 15 <= ts.hour <= 23:
                                    group_df.at[idx, "remark"] = "早退"
                        else:
                            if not (0 <= ts.hour <= 1):
                                if ts_time < end_time:
                                    group_df.at[idx, "remark"] = "早退"

            group_df = group_df.sort_values(["name", "timestamp"], na_position="last")
            slim_df = group_df[["name", "timestamp", "keyword", "shift", "remark"]].copy()
            slim_df.columns = ["姓名", "打卡时间", "关键词", "班次", "备注"]

            slim_df["打卡时间"] = slim_df["打卡时间"].apply(lambda x: x.strftime("%H:%M:%S") if pd.notna(x) else "")
            slim_df["班次"] = slim_df["班次"].apply(format_shift)

            sheet_name = day[:31]
            sheet = writer.book.create_sheet(sheet_name)
            headers = ["姓名", "打卡时间", "关键词", "班次", "备注"]
            sheet.append(headers)

            for user, user_df in slim_df.groupby("姓名"):
                for _, row in user_df.iterrows():
                    sheet.append(list(row))
                sheet.append([None] * len(headers))

            sheet_written = True

        if not sheet_written:
            pd.DataFrame(columns=["姓名", "打卡时间", "关键词", "班次", "备注"]).to_excel(writer, sheet_name="空表", index=False)

    # ======================== 样式处理 ========================
    wb = load_workbook(excel_path)
    red_fill = PatternFill(start_color="ffc8c8", end_color="ffc8c8", fill_type="solid")
    yellow_fill = PatternFill(start_color="fff1c8", end_color="fff1c8", fill_type="solid")
    blue_fill_light = PatternFill(start_color="c8eaff", end_color="c8eaff", fill_type="solid")
    purple_fill_light = PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    from itertools import cycle
    user_fills = cycle([
        PatternFill(start_color="f9f9f9", end_color="f9f9f9", fill_type="solid"),
        PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid"),
    ])

    for sheet in wb.worksheets:
        if sheet.title == "统计":
            continue
        current_user = None
        current_fill = next(user_fills)
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                continue
            name_val = row[0].value
            remark_val = str(row[4].value or "")
            if name_val != current_user:
                current_fill = next(user_fills)
                current_user = name_val
            for cell in row:
                cell.fill = current_fill
            if "迟到" in remark_val or "早退" in remark_val:
                for cell in row[1:]:
                    cell.fill = red_fill
            elif "补卡" in remark_val:
                for cell in row[1:]:
                    cell.fill = yellow_fill
            elif "休息/缺勤" in remark_val:
                for cell in row[1:]:
                    cell.fill = blue_fill_light
            elif "未打下班卡" in remark_val:
                for cell in row[1:]:
                    cell.fill = purple_fill_light

        # 合并姓名列
        name_col = 1
        merge_start = None
        prev_name = None
        for row_idx in range(2, sheet.max_row + 1):
            cell_val = sheet.cell(row=row_idx, column=name_col).value
            if cell_val != prev_name:
                if merge_start and row_idx - merge_start > 1:
                    sheet.merge_cells(
                        start_row=merge_start, start_column=name_col,
                        end_row=row_idx - 1, end_column=name_col
                    )
                merge_start = row_idx
                prev_name = cell_val
        if merge_start and sheet.max_row - merge_start >= 1:
            sheet.merge_cells(
                start_row=merge_start, start_column=name_col,
                end_row=sheet.max_row, end_column=name_col
            )

# ======================== 统计表（终极修正版） ========================
    stats = {u: {"休息/缺勤": 0, "迟到/早退": 0, "补卡": 0, "未打下班卡": 0} for u in all_user_names}
    
    for sheet in wb.worksheets:
        if sheet.title == "统计":
            continue
        df_sheet = pd.DataFrame(sheet.values)
        if df_sheet.empty or len(df_sheet.columns) < 5:
            continue
        df_sheet.columns = ["姓名", "打卡时间", "关键词", "班次", "备注"]
    
        # 补齐姓名
        last_name = None
        for i in range(len(df_sheet)):
            if pd.notna(df_sheet.at[i, "姓名"]):
                last_name = df_sheet.at[i, "姓名"]
            elif last_name:
                df_sheet.at[i, "姓名"] = last_name
    
        df_sheet["备注"] = df_sheet["备注"].astype(str).fillna("")
    
        # 分组统计
        for name, g in df_sheet.groupby("姓名"):
            if not name or name not in stats:
                continue
    
            stats[name]["补卡"] += int(g["备注"].str.count("补卡").sum())
            stats[name]["迟到/早退"] += int(g["备注"].str.count("迟到").sum() +
                                         g["备注"].str.count("早退").sum())
            stats[name]["休息/缺勤"] += int(g["备注"].str.count("休息/缺勤").sum())
            stats[name]["未打下班卡"] += int(g["备注"].str.count("未打下班卡").sum())
    
    # 转换为 DataFrame（去掉正常列）
    summary_df = pd.DataFrame([
        {
            "姓名": u,
            **v,
            "异常总数": v["迟到/早退"] + v["补卡"] + v["未打下班卡"]
        }
        for u, v in stats.items()
    ])
    summary_df = summary_df[["姓名", "休息/缺勤", "迟到/早退", "补卡", "未打下班卡", "异常总数"]]
    
    # 写入 Excel
    if "统计" in [s.title for s in wb.worksheets]:
        del wb["统计"]
    
    stats_sheet = wb.create_sheet("统计", 0)
    headers = ["姓名", "休息/缺勤", "迟到/早退", "补卡", "未打下班卡", "异常总数"]
    for r_idx, row in enumerate([headers] + summary_df.values.tolist(), 1):
        for c_idx, value in enumerate(row, 1):
            stats_sheet.cell(row=r_idx, column=c_idx, value=value)

    # 样式
    stats_sheet.freeze_panes = "A2"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    highlight_fill = PatternFill(start_color="FFF8B0", end_color="FFF8B0", fill_type="solid")

    stats_sheet.auto_filter.ref = stats_sheet.dimensions

    for cell in stats_sheet[1]:
        cell.font = header_font
        cell.alignment = center_align

    light_red_fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
    
    for row in stats_sheet.iter_rows(min_row=2):
        try:
            # 如果休息/缺勤 > 4，则标淡红色（第 3 列）
            rest_days = int(row[2].value or 0)
            if rest_days > 4:
                row[2].fill = light_red_fill
    
            # 如果异常总数 > 2，则标淡红色（最后一列）
            abnormal_total = int(row[-1].value or 0)
            if abnormal_total > 2:
                row[-1].fill = light_red_fill
        except ValueError:
            pass

    # ======================== 说明文字 ========================
    desc_text = (
        "【休息/缺勤：没有打卡记录的天数】\n"
        "【异常总数：迟到/早退+补卡+未打下班卡】"
    )
    
    start_row = summary_df.shape[0] + 3  # 表格最后一行 + 空一行
    end_row = start_row + 2  # 三行高度
    
    # 合并三行七列
    stats_sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=6)
    cell = stats_sheet.cell(row=start_row, column=1, value=desc_text)
    
    # 居中对齐 + 自动换行
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 标准黄色底色
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    cell.fill = yellow_fill
    
    # 加粗字体（黑色）
    cell.font = Font(bold=True, color="000000")

    # ======================== 列宽/边框 + 自动筛选 ========================
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions  

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in sheet.columns:
            col_letter = col[0].column_letter
            max_length = max((19 if isinstance(cell.value, datetime) else len(str(cell.value or "")) for cell in col))
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            sheet.column_dimensions[col_letter].width = min(max_length + 8, 30)

    wb.save(excel_path)
    logging.info(f"✅ Excel 导出完成: {excel_path}")
    return excel_path


def export_user_excel(user_name: str, start_datetime: datetime, end_datetime: datetime):
    df = _fetch_data(start_datetime, end_datetime)
    if df.empty:
        logging.warning(f"⚠️ 指定日期内没有 {user_name} 的数据")
        return None

    # 只筛选该用户
    df = df[df["name"] == user_name]
    if df.empty:
        logging.warning(f"⚠️ {user_name} 在指定日期没有考勤记录")
        return None

    # ======================== 时间处理 ========================
    if pd.api.types.is_datetime64_any_dtype(df["timestamp"]):
        try:
            df["timestamp"] = df["timestamp"].dt.tz_localize(None)
        except AttributeError:
            pass

    df["日期"] = df["timestamp"].dt.strftime("%Y-%m-%d")

    def format_shift(shift):
        if pd.isna(shift):
            return shift
        shift_text = str(shift)
        if re.search(r'（\d{2}:\d{2}-\d{2}:\d{2}）', shift_text):
            return shift_text
        shift_name = shift_text.split("（")[0]
        if shift_name in get_shift_times_short():
            start, end = get_shift_times_short()[shift_name]
            return f"{shift_text}（{start.strftime('%H:%M')}-{end.strftime('%H:%M')}）"
        return shift_text

    # ======================== remark 标注逻辑 ========================
    if "remark" not in df.columns:
        df["remark"] = ""

    for idx, row in df.iterrows():
        shift_val = row["shift"]
        keyword = row["keyword"]
        ts = row["timestamp"]

        if not shift_val or pd.isna(ts):
            continue

        shift_text = str(shift_val).strip()
        shift_name = re.split(r'[（(]', shift_text)[0]

        if "补卡" in shift_text:
            df.at[idx, "remark"] = "补卡"
            continue

        if shift_name in get_shift_times_short():
            start_time, end_time = get_shift_times_short()[shift_name]
            ts_time = ts.time()

            if keyword == "#上班打卡" and ts_time > start_time:
                df.at[idx, "remark"] = "迟到"
            elif keyword == "#下班打卡":
                if shift_name == "I班":
                    if not (ts.hour == 0):
                        if 15 <= ts.hour <= 23:
                            df.at[idx, "remark"] = "早退"
                else:
                    if not (0 <= ts.hour <= 1):
                        if ts_time < end_time:
                            df.at[idx, "remark"] = "早退"

    # ======================== 处理 I 班跨日下班卡 ========================
    i_shift_mask = (
        (df["keyword"] == "#下班打卡") &
        (df["shift"].notna()) &
        (df["shift"].astype(str).str.startswith("I班")) &
        (df["timestamp"].dt.hour < 6)
    )
    cross_df = df[i_shift_mask].copy()
    df = df[~i_shift_mask]
    cross_df["remark"] = cross_df.get("remark", "") + "（次日）"
    cross_df["日期"] = (cross_df["timestamp"] - pd.Timedelta(days=1)).dt.strftime("%Y-%m-%d")
    df = pd.concat([df, cross_df], ignore_index=True)

    # ======================== 补齐休息/缺勤 ========================
    all_dates = pd.date_range(start_datetime.date(), (end_datetime - timedelta(seconds=1)).date(), freq="D")
    existing_dates = set(df["日期"].unique())
    missing_dates = [d.strftime("%Y-%m-%d") for d in all_dates if d.strftime("%Y-%m-%d") not in existing_dates]

    if missing_dates:
        missed_df = pd.DataFrame({
            "日期": missing_dates,
            "name": user_name,
            "timestamp": pd.NaT,
            "keyword": None,
            "shift": None,
            "remark": "休息/缺勤"
        })
        df = pd.concat([df, missed_df], ignore_index=True)

    # ======================== 补齐未打下班卡 ========================
    unclosed_rows = []
    for day, g in df.groupby("日期"):
        has_up = g["keyword"].eq("#上班打卡").any()
        has_down = g["keyword"].eq("#下班打卡").any()
        if has_up and not has_down:
            unclosed_rows.append({
                "日期": day,
                "name": user_name,
                "timestamp": pd.NaT,
                "keyword": "#下班打卡",
                "shift": None,
                "remark": "未打下班卡"
            })
    if unclosed_rows:
        df = pd.concat([df, pd.DataFrame(unclosed_rows)], ignore_index=True)

    # ======================== 整理数据表 ========================
    slim_df = df[["日期", "name", "timestamp", "keyword", "shift", "remark"]].copy()
    slim_df.columns = ["日期", "姓名", "打卡时间", "关键词", "班次", "备注"]

    slim_df["打卡时间"] = slim_df["打卡时间"].apply(lambda x: x.strftime("%H:%M:%S") if pd.notna(x) else "")
    slim_df["班次"] = slim_df["班次"].apply(format_shift)

    keyword_order = {"#上班打卡": 0, "#下班打卡": 1, None: 2}
    slim_df["kw_order"] = slim_df["关键词"].map(keyword_order).fillna(9)
    slim_df = slim_df.sort_values(["日期", "姓名", "班次", "kw_order", "打卡时间"]).drop(columns=["kw_order"])

    # ======================== 导出 Excel ========================
    start_str = start_datetime.strftime("%Y-%m-%d")
    end_str = (end_datetime - pd.Timedelta(seconds=1)).strftime("%Y-%m-%d")
    export_dir = os.path.join(DATA_DIR, f"user_excel_{start_str}_{end_str}")
    os.makedirs(export_dir, exist_ok=True)
    file_path = os.path.join(export_dir, f"{user_name}_考勤详情.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{user_name}考勤详情"

    headers = ["日期", "姓名", "打卡时间", "关键词", "班次", "备注"]
    ws.append(headers)

    for _, row in slim_df.iterrows():
        ws.append(list(row))

    # ======================== 样式处理 ========================
    red_fill = PatternFill(start_color="ffc8c8", end_color="ffc8c8", fill_type="solid")
    yellow_fill = PatternFill(start_color="fff1c8", end_color="fff1c8", fill_type="solid")
    blue_fill_light = PatternFill(start_color="c8eaff", end_color="c8eaff", fill_type="solid")
    purple_fill = PatternFill(start_color="e6ccff", end_color="e6ccff", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    from itertools import cycle
    user_fills = cycle([
        PatternFill(start_color="f9f9f9", end_color="f9f9f9", fill_type="solid"),
        PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid"),
    ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_fill = next(user_fills)
    prev_date = None
    for row in ws.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            continue
        date_val = row[0].value
        remark_val = str(row[5].value or "")

        if date_val != prev_date:
            current_fill = next(user_fills)
            prev_date = date_val

        for cell in row:
            cell.fill = current_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        if "迟到" in remark_val or "早退" in remark_val:
            for cell in row[2:]:
                cell.fill = red_fill
        elif "补卡" in remark_val:
            for cell in row[2:]:
                cell.fill = yellow_fill
        elif "休息/缺勤" in remark_val:
            for cell in row[2:]:
                cell.fill = blue_fill_light
        elif "未打下班卡" in remark_val:
            for cell in row[2:]:
                cell.fill = purple_fill

    # 列宽自适应
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = min(max_length + 6, 30)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(file_path)
    logging.info(f"✅ 已导出用户 {user_name} 的考勤详情：{file_path}")
    return file_path
